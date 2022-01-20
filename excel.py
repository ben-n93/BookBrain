from openpyxl import Workbook
import book_brain
import json
wb = Workbook()
ws = wb.active
ws.title = "Book entries"

ws['A1'] = 'ID'
ws['B1'] = 'Date started:'
ws['C1'] = 'Date finished:'
ws['D1'] = 'Book title:'
ws['E1'] = 'Author title:'
ws['F1'] = 'Genre:'
ws['G1'] = 'Comments:'

entries = {}

entries_file = 'data/user_data.json'

with open(entries_file) as f:
	entries = json.load(f)

row_count = 0

temp_entries = []

book_title = 'Dune '
book_title = book_title.rstrip()

for dictionary in entries.values():
	if book_title in dictionary.values():
		temp_entries.append(dictionary['ID'])

print(temp_entries)

#wb.save('BookBrain.xlsx')